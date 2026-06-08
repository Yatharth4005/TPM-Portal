# seed_workstations.py
import os
import django
import datetime
import re
import openpyxl

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'jspl_tpm.settings')
django.setup()

from tpm.models import Department, Workstation, WorkstationKPI, WorkstationValue

SHEET_MAPPING = {
    'PlateMill': 'PM',
    'SPM': 'SPM',
    'Rail Mill': 'RM',
    'BF-1': 'BF1',
    'BF-2': 'BF2',
    'SMS-2': 'SMS2',
    'SMS-3': 'SMS3',
    'RMH-1': 'RMHS1',
    'RMH-3': 'RMHS3'
}

LOWER_IS_BETTER_KEYWORDS = [
    'breakdown', 'mttr', 'total losses', 'rejection', 'rework',
    'response time', 'repetitive breakdown', 'customer complaint',
    'fatal', 'lti', 'mti', 'near miss', 'accident', 'incident',
    'low illuminated area identified', 'dust level', 'noise level',
    'delay', 'leakage', 'contamination', 'tat', 'vibration'
]

def clean_float(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    
    val_str = str(val).strip().lower()
    if val_str in ('na', 'nil', 'nil', 'none', 'zero'):
        if val_str == 'zero':
            return 0.0
        return None
        
    # Extract first float/int match using regex
    match = re.search(r'[-+]?\d*\.\d+|\d+', val_str)
    if match:
        return float(match.group())
    return None

def parse_date(date_val):
    if not date_val:
        return None
    if isinstance(date_val, datetime.date):
        return date_val
    if isinstance(date_val, datetime.datetime):
        return date_val.date()
    
    # Try parsing string format
    date_str = str(date_val).strip()
    for fmt in ('%d.%m.%Y', '%d.%m.%y', '%Y-%m-%d', '%Y-%m-%d %H:%M:%S'):
        try:
            return datetime.datetime.strptime(date_str, fmt).date()
        except ValueError:
            pass
    return None

def main():
    excel_path = "Workstation KPI Trackers.xlsx"
    if not os.path.exists(excel_path):
        print(f"Error: {excel_path} not found.")
        return

    print("Loading workbook...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    total_workstations_created = 0
    total_workstations_matched = 0
    total_kpis_created = 0
    total_kpis_matched = 0
    total_values_created = 0
    total_values_updated = 0

    for sheet_name in wb.sheetnames:
        dept_code = SHEET_MAPPING.get(sheet_name)
        if not dept_code:
            continue

        try:
            dept = Department.objects.get(code=dept_code)
        except Department.DoesNotExist:
            continue

        print(f"\nProcessing sheet '{sheet_name}' for department '{dept.name}' ({dept_code})...")
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            print("  No rows found in worksheet.")
            continue

        # Find header row
        header = None
        header_idx = 0
        for idx, row in enumerate(rows):
            if row and any(str(x).strip().lower() in ['s. no.', 's.no.', 'sno'] for x in row if x is not None):
                header = [str(x).strip() if x is not None else f"Col{i}" for i, x in enumerate(row)]
                header_idx = idx
                break
        
        if not header:
            header = [str(x).strip() if x is not None else f"Col{i}" for i, x in enumerate(rows[0])]
            header_idx = 0

        print(f"  Header row found at index {header_idx}: {header[:10]}")

        # Parse date headers for columns 10 onwards
        month_columns = []
        for col_idx in range(10, len(header)):
            header_val = header[col_idx]
            parsed_date = parse_date(header_val)
            if parsed_date:
                month_columns.append((col_idx, parsed_date.month, parsed_date.year))

        print(f"  Month columns detected: {[(idx, f'{m}/{y}') for idx, m, y in month_columns]}")

        data_rows_processed = 0
        for r_idx, row in enumerate(rows[header_idx + 1:]):
            if not row or all(x is None for x in row):
                continue

            workstation_name = row[2] if len(row) > 2 else None
            leader = row[3] if len(row) > 3 else None
            inception_val = row[4] if len(row) > 4 else None
            kpi_name = row[5] if len(row) > 5 else None
            uom = row[6] if len(row) > 6 else None
            goodness_val = row[7] if len(row) > 7 else None
            baseline_val = row[8] if len(row) > 8 else None
            commitment_val = row[9] if len(row) > 9 else None

            if not workstation_name or not kpi_name:
                # Debug why skipped
                if any(x is not None for x in row):
                    pass
                continue

            data_rows_processed += 1
            workstation_name = str(workstation_name).strip()
            kpi_name = str(kpi_name).strip()
            leader = str(leader).strip() if leader else "N/A"
            uom = str(uom).strip() if uom else ""

            # Parse inception date
            inception_date = parse_date(inception_val)
            if not inception_date:
                inception_date = datetime.date(2025, 4, 1)

            # Get or create workstation
            workstation, ws_created = Workstation.objects.get_or_create(
                department=dept,
                name=workstation_name,
                defaults={'leader': leader, 'inception_date': inception_date}
            )
            if ws_created:
                total_workstations_created += 1
            else:
                total_workstations_matched += 1

            # Determine goodness indicator
            goodness_indicator = WorkstationKPI.GoodnessIndicator.HIGHER
            lower_name = kpi_name.lower()
            if any(kw in lower_name for kw in LOWER_IS_BETTER_KEYWORDS):
                goodness_indicator = WorkstationKPI.GoodnessIndicator.LOWER
            elif commitment_val and any(x in str(commitment_val).lower() for x in ('<', 'lower', 'min')):
                goodness_indicator = WorkstationKPI.GoodnessIndicator.LOWER
            elif goodness_val and any(x in str(goodness_val).lower() for x in ('lower', 'min')):
                goodness_indicator = WorkstationKPI.GoodnessIndicator.LOWER

            # Clean baseline & commitment
            baseline = clean_float(baseline_val)
            commitment = clean_float(commitment_val)

            # Get or create workstation KPI
            kpi, kpi_created = WorkstationKPI.objects.get_or_create(
                workstation=workstation,
                kpi_name=kpi_name,
                defaults={
                    'uom': uom,
                    'goodness_indicator': goodness_indicator,
                    'baseline': baseline,
                    'commitment': commitment
                }
            )
            if kpi_created:
                total_kpis_created += 1
            else:
                total_kpis_matched += 1

            # Parse monthly values
            for col_idx, month, year in month_columns:
                if col_idx < len(row):
                    cell_val = row[col_idx]
                    actual = clean_float(cell_val)
                    if actual is not None:
                        val, val_created = WorkstationValue.objects.update_or_create(
                            workstation_kpi=kpi,
                            month=month,
                            year=year,
                            defaults={'actual': actual}
                        )
                        if val_created:
                            total_values_created += 1
                        else:
                            total_values_updated += 1
        
        print(f"  Processed {data_rows_processed} data rows for sheet '{sheet_name}'.")

    print("\n" + "="*40)
    print("Seeding Summary:")
    print(f"Workstations: {total_workstations_created} created, {total_workstations_matched} matched.")
    print(f"KPIs:         {total_kpis_created} created, {total_kpis_matched} matched.")
    print(f"Values:       {total_values_created} created, {total_values_updated} updated/matched.")
    print("="*40)

if __name__ == "__main__":
    main()

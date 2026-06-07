# tpm/management/commands/import_wkpi.py

import datetime
import re
import openpyxl
from django.core.management.base import BaseCommand
from tpm.models import Department, Workstation, WorkstationKPI, WorkstationValue

class Command(BaseCommand):
    help = 'Imports Workstation KPIs and monthly values from Workstation KPI Trackers.xlsx'

    def clean_num(self, val):
        if val is None:
            return None
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val).strip().lower()
        if s in ['zero', 'nil', 'nil ', 'nill']:
            return 0.0
        if s in ['na', 'n/a', '']:
            return None
        # Try to extract decimal number
        m = re.search(r'([0-9]+(?:\.[0-9]+)?)', s)
        if m:
            return float(m.group(1))
        return None

    def parse_date(self, val):
        if val is None:
            return None
        if isinstance(val, datetime.date):
            return val
        if isinstance(val, datetime.datetime):
            return val.date()
        s = str(val).strip()
        if not s:
            return None
        # Try format patterns: dd.mm.yyyy, dd.mm.yy, yyyy-mm-dd
        for fmt in ('%d.%m.%Y', '%d.%m.%y', '%Y-%m-%d'):
            try:
                return datetime.datetime.strptime(s, fmt).date()
            except ValueError:
                pass
        return None

    def handle(self, *args, **kwargs):
        self.stdout.write('Loading workbook e:/TPM Portal/Workstation KPI Trackers.xlsx...')
        try:
            wb = openpyxl.load_workbook('e:/TPM Portal/Workstation KPI Trackers.xlsx', data_only=True)
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Failed to load Excel workbook: {e}'))
            return

        sheet_mapping = {
            'PlateMill': 'PM',
            'SPM': 'SPM',
            'Rail Mill': 'RM',
            'BF-2': 'BF2',
            'BF-1': 'BF1',
            'SMS-2': 'SMS2',
            'SMS-3': 'SMS3',
            'RMH-1': 'RMHS1',
            'RMH-3': 'RMHS3'
        }

        total_workstations = 0
        total_kpis = 0
        total_values = 0

        # Clean existing workstation data for mapped departments to avoid duplication
        for s_name, code in sheet_mapping.items():
            dept = Department.objects.filter(code=code).first()
            if dept:
                # Cascade deletes Workstations, WorkstationKPIs, and WorkstationValues
                Workstation.objects.filter(department=dept).delete()
                self.stdout.write(f'Cleared existing Workstation KPI entries for {dept.name} ({code}).')

        for sheet_name, dept_code in sheet_mapping.items():
            if sheet_name not in wb.sheetnames:
                self.stdout.write(self.style.WARNING(f'Sheet {sheet_name} not found in Excel workbook, skipping.'))
                continue

            dept = Department.objects.filter(code=dept_code).first()
            if not dept:
                self.stdout.write(self.style.ERROR(f'Department with code {dept_code} not found, skipping sheet {sheet_name}.'))
                continue

            ws = wb[sheet_name]
            self.stdout.write(f'Processing sheet: {sheet_name} ({dept.name})...')
            
            headers = [cell.value for cell in ws[1]]
            
            row_count = 0
            for r in range(2, ws.max_row + 1):
                # Columns: 0=S.No, 1=Plant, 2=Workstation, 3=Leader, 4=WS Inception Date, 5=KPI, 6=UoM, 7=Goodness Indicator, 8=Baseline, 9=Commitment
                ws_name = ws.cell(row=r, column=3).value
                leader = ws.cell(row=r, column=4).value
                inception_raw = ws.cell(row=r, column=5).value
                kpi_name = ws.cell(row=r, column=6).value
                uom = ws.cell(row=r, column=7).value
                baseline_raw = ws.cell(row=r, column=9).value
                commitment_raw = ws.cell(row=r, column=10).value

                if not ws_name or not kpi_name:
                    continue

                row_count += 1
                
                # Parse inception date
                inception_date = self.parse_date(inception_raw)
                if not inception_date:
                    inception_date = datetime.date(2025, 4, 1) # Fallback to start of tracking

                # 1. Save Workstation
                workstation, ws_created = Workstation.objects.get_or_create(
                    department=dept,
                    name=str(ws_name).strip(),
                    defaults={
                        'leader': str(leader).strip() if leader else 'N/A',
                        'inception_date': inception_date
                    }
                )
                if ws_created:
                    total_workstations += 1

                # 2. Determine Goodness Indicator based on keywords
                lower_keywords = [
                    'breakdown', 'delay', 'rejection', 'rework', 'absenteeism', 
                    'accident', 'harm', 'failure', 'consumption', 'waste', 
                    'dust', 'noise', 'incident', 'time', 'loss'
                ]
                
                # Default is HIGHER. If lower-is-better keyword in KPI, set to LOWER.
                goodness = 'HIGHER'
                kpi_lower = any(kw in str(kpi_name).lower() for kw in lower_keywords)
                if kpi_lower:
                    goodness = 'LOWER'

                # 3. Save Workstation KPI
                baseline = self.clean_num(baseline_raw)
                commitment = self.clean_num(commitment_raw)

                wk_kpi, kpi_created = WorkstationKPI.objects.get_or_create(
                    workstation=workstation,
                    kpi_name=str(kpi_name).strip(),
                    defaults={
                        'uom': str(uom).strip() if uom else '',
                        'goodness_indicator': goodness,
                        'baseline': baseline,
                        'commitment': commitment
                    }
                )
                total_kpis += 1

                # 4. Save Monthly Values
                for col_idx in range(11, len(headers) + 1):
                    h_val = headers[col_idx-1]
                    if h_val is None:
                        continue

                    # Parse month/year from date header
                    if isinstance(h_val, (datetime.date, datetime.datetime)):
                        month = h_val.month
                        year = h_val.year
                    elif isinstance(h_val, str):
                        # try string split e.g. 2025-04 or 01.04.2025
                        m = re.match(r'(\d{4})-(\d{2})', h_val)
                        if m:
                            year = int(m.group(1))
                            month = int(m.group(2))
                        else:
                            # Try other parsing or skip
                            continue
                    else:
                        continue

                    cell_val = ws.cell(row=r, column=col_idx).value
                    actual = self.clean_num(cell_val)

                    if actual is not None:
                        # Save value in DB
                        WorkstationValue.objects.update_or_create(
                            workstation_kpi=wk_kpi,
                            month=month,
                            year=year,
                            defaults={'actual': actual}
                        )
                        total_values += 1

            self.stdout.write(f'  Sheet {sheet_name} imported: {row_count} KPI rows.')

        self.stdout.write(self.style.SUCCESS(
            f'Import completed successfully!\n'
            f'- Seeded Workstations: {total_workstations}\n'
            f'- Seeded KPIs: {total_kpis}\n'
            f'- Seeded Values: {total_values}'
        ))

import openpyxl
import json

wb = openpyxl.load_workbook("Workstation KPI Trackers.xlsx", data_only=True)
sheets = wb.sheetnames

all_data = {}

for sheet in sheets:
    if sheet == 'Sheet1': # Sheet1 seems to be a duplicate or combined sheet, let's look closely later
        continue
    ws = wb[sheet]
    # Find headers
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        continue
    
    header = None
    header_idx = 0
    for idx, row in enumerate(rows):
        if row and any(str(x).strip().lower() in ['s. no.', 's.no.', 'sno'] for x in row if x is not None):
            header = [str(x).strip() if x is not None else f"Col{i}" for i, x in enumerate(row)]
            header_idx = idx
            break
            
    if not header:
        # Fallback to row 0 if no header found
        header = [str(x).strip() if x is not None else f"Col{i}" for i, x in enumerate(rows[0])]
        header_idx = 0
        
    print(f"Sheet: {sheet}, Header row: {header_idx}, Fields: {header[:12]}")
    
    sheet_records = []
    for row in rows[header_idx + 1:]:
        if not row or all(x is None for x in row):
            continue
        # Check if s. no. is empty and workstation/kpi are empty
        s_no = row[0] if len(row) > 0 else None
        plant = row[1] if len(row) > 1 else None
        workstation = row[2] if len(row) > 2 else None
        leader = row[3] if len(row) > 3 else None
        inception = row[4] if len(row) > 4 else None
        kpi = row[5] if len(row) > 5 else None
        uom = row[6] if len(row) > 6 else None
        goodness = row[7] if len(row) > 7 else None
        baseline = row[8] if len(row) > 8 else None
        commitment = row[9] if len(row) > 9 else None
        
        # If both workstation and kpi are None, skip
        if not workstation and not kpi:
            continue
            
        sheet_records.append({
            "s_no": s_no,
            "plant": plant,
            "workstation": workstation,
            "leader": leader,
            "inception": str(inception) if inception else None,
            "kpi": kpi,
            "uom": uom,
            "goodness": goodness,
            "baseline": baseline,
            "commitment": commitment
        })
    all_data[sheet] = sheet_records

with open("extracted_kpis.json", "w") as f:
    json.dump(all_data, f, indent=4)

print("Extraction completed!")

import openpyxl

wb = openpyxl.load_workbook("Workstation KPI Trackers.xlsx", data_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=================== Sheet: {sheet_name} ===================")
    print(f"Dimensions: {ws.dimensions}")
    
    # Let's inspect the first 15 rows and first 10 columns of each sheet to get the layout
    rows = list(ws.iter_rows(max_row=25, max_col=15, values_only=True))
    for r_idx, row in enumerate(rows):
        # Only print rows that have at least one non-None value
        if any(v is not None for v in row):
            row_str = " | ".join(str(v)[:20] if v is not None else "" for v in row)
            print(f"Row {r_idx+1:02d}: {row_str}")

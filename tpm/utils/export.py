# tpm/utils/export.py

import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

from tpm.models import PillarEntry, KPIValue, WorkstationValue, Workstation
from tpm.utils.kpi_definitions import KPI_DEFINITIONS
from tpm.utils.calculations import compute_achievement, get_date_range_q, aggregate_kpi_actual

def generate_pillar_excel(dept, from_month, from_year, to_month, to_year, filter_type):
    """Generates an Excel workbook with sheets for Summary and each active Pillar"""
    wb = Workbook()
    
    months_map = dict([
        (1, 'Jan'), (2, 'Feb'), (3, 'Mar'), (4, 'Apr'),
        (5, 'May'), (6, 'Jun'), (7, 'Jul'), (8, 'Aug'),
        (9, 'Sep'), (10, 'Oct'), (11, 'Nov'), (12, 'Dec')
    ])
    if filter_type == 'range':
        period_label = f"{months_map.get(from_month)} {from_year} - {months_map.get(to_month)} {to_year}"
    else:
        months_full = [
            (1, 'January'), (2, 'February'), (3, 'March'), (4, 'April'),
            (5, 'May'), (6, 'June'), (7, 'July'), (8, 'August'),
            (9, 'September'), (10, 'October'), (11, 'November'), (12, 'December')
        ]
        period_label = f"{dict(months_full).get(from_month)} {from_year}"

    # 1. Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Summary Overview"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws_summary.merge_cells('A1:E2')
    title_cell = ws_summary['A1']
    title_cell.value = f"JINDAL STEEL & POWER LTD — TPM PORTAL"
    title_cell.font = Font(name='Segoe UI', size=16, bold=True, color='003478')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws_summary['A3'] = f"Department: {dept.name} ({dept.code})"
    ws_summary['A3'].font = Font(name='Segoe UI', size=11, bold=True)
    ws_summary['A4'] = f"Report Period: {period_label}"
    ws_summary['A4'].font = Font(name='Segoe UI', size=11, italic=True)
    
    headers = ["Pillar Code", "Pillar Name", "Total KPIs", "On Track (>=90%)", "At Risk (75-89%)", "Behind (<75%)"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws_summary.cell(row=6, column=col_idx)
        cell.value = h
        cell.font = Font(name='Segoe UI', bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='003478', end_color='003478', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        
    thin_border = Border(
        left=Side(style='thin', color='D1DCF0'),
        right=Side(style='thin', color='D1DCF0'),
        top=Side(style='thin', color='D1DCF0'),
        bottom=Side(style='thin', color='D1DCF0')
    )
    
    pillars_meta = [
        ('KK', 'Kobetsu Kaizen'),
        ('JH', 'Jishu Hozen'),
        ('PM', 'Planned Maintenance'),
        ('QM', 'Quality Maintenance'),
        ('ET', 'Education & Training'),
        ('DM', 'Initial Flow Control / Design & Management'),
        ('SHE', 'Safety, Health & Environment'),
        ('OTPM', 'Office TPM')
    ]
    
    row_idx = 7
    for code, name in pillars_meta:
        definitions = KPI_DEFINITIONS.get(code, [])
        total = len(definitions)
        on_track = 0
        at_risk = 0
        behind = 0
        has_any_data = False
        
        for d in definitions:
            kpi_values = KPIValue.objects.filter(
                get_date_range_q(prefix='pillar_entry__', from_month=from_month, from_year=from_year, to_month=to_month, to_year=to_year),
                pillar_entry__department=dept,
                pillar_entry__pillar=code,
                sl_no=d['sl_no']
            )
            if kpi_values.exists():
                has_any_data = True
                actual, target, benchmark = aggregate_kpi_actual(kpi_values, d['uom'], d['name'])
                if actual is not None and target is not None:
                    ach = compute_achievement(actual, target, d['name'])
                    if ach >= 90:
                        on_track += 1
                    elif ach >= 75:
                        at_risk += 1
                    else:
                        behind += 1
                        
        ws_summary.cell(row=row_idx, column=1, value=code)
        ws_summary.cell(row=row_idx, column=2, value=name)
        ws_summary.cell(row=row_idx, column=3, value=total if has_any_data else "N/A")
        ws_summary.cell(row=row_idx, column=4, value=on_track if has_any_data else "N/A")
        ws_summary.cell(row=row_idx, column=5, value=at_risk if has_any_data else "N/A")
        ws_summary.cell(row=row_idx, column=6, value=behind if has_any_data else "N/A")
        
        for col in range(1, 7):
            c = ws_summary.cell(row=row_idx, column=col)
            c.border = thin_border
            c.font = Font(name='Segoe UI')
            if col > 2:
                c.alignment = Alignment(horizontal='center')
        
        row_idx += 1
        
    # Autofit columns
    for col in ws_summary.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_summary.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # 2. Add individual Pillar Sheets
    for code, name in pillars_meta:
        ws = wb.create_sheet(title=code)
        ws.views.sheetView[0].showGridLines = True
        
        ws.cell(row=1, column=1, value=f"{code} — {name} Report ({period_label})").font = Font(name='Segoe UI', size=14, bold=True, color='003478')
        
        headers = ["Sl No", "KPI Name", "UOM", "Benchmark", "Target", "Actual", "Achievement %", "Remarks"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = h
            cell.font = Font(name='Segoe UI', bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='0057A8', end_color='0057A8', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            
        definitions = KPI_DEFINITIONS.get(code, [])
        
        p_row = 4
        for d in definitions:
            kpi_values = KPIValue.objects.filter(
                get_date_range_q(prefix='pillar_entry__', from_month=from_month, from_year=from_year, to_month=to_month, to_year=to_year),
                pillar_entry__department=dept,
                pillar_entry__pillar=code,
                sl_no=d['sl_no']
            )
            
            if kpi_values.exists():
                actual, target, benchmark = aggregate_kpi_actual(kpi_values, d['uom'], d['name'])
                remarks_list = [v.remarks.strip() for v in kpi_values if v.remarks.strip()]
                remarks = " | ".join(remarks_list) if remarks_list else ""
            else:
                actual = None
                target = d['target']
                benchmark = d['benchmark']
                remarks = ''
            
            achievement = ""
            if actual is not None and target is not None:
                ach = compute_achievement(actual, target, d['name'])
                achievement = f"{round(ach, 1)}%"
                
            ws.cell(row=p_row, column=1, value=d['sl_no']).alignment = Alignment(horizontal='center')
            ws.cell(row=p_row, column=2, value=d['name'])
            ws.cell(row=p_row, column=3, value=d['uom']).alignment = Alignment(horizontal='center')
            
            ws.cell(row=p_row, column=4, value=benchmark if benchmark is not None else "—").alignment = Alignment(horizontal='right')
            ws.cell(row=p_row, column=5, value=target if target is not None else "—").alignment = Alignment(horizontal='right')
            ws.cell(row=p_row, column=6, value=actual if actual is not None else "—").alignment = Alignment(horizontal='right')
            ws.cell(row=p_row, column=7, value=achievement).alignment = Alignment(horizontal='right')
            ws.cell(row=p_row, column=8, value=remarks)
            
            # Format row
            for col in range(1, 9):
                c = ws.cell(row=p_row, column=col)
                c.border = thin_border
                c.font = Font(name='Segoe UI')
                
            p_row += 1
            
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)
            
    # Save to IO stream
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_pillar_pdf(dept, from_month, from_year, to_month, to_year, filter_type):
    """Generates a professional PDF monthly report for the department using ReportLab"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40
    )
    
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        name='JSPLTitle',
        fontName='Helvetica-Bold',
        fontSize=18,
        leading=22,
        textColor=colors.HexColor('#003478'),
        alignment=1 # Center
    )
    
    subtitle_style = ParagraphStyle(
        name='JSPLSubtitle',
        fontName='Helvetica-Oblique',
        fontSize=11,
        leading=14,
        textColor=colors.HexColor('#6B7A99'),
        alignment=1
    )
    
    header_style = ParagraphStyle(
        name='JSPLHeader',
        fontName='Helvetica-Bold',
        fontSize=14,
        leading=18,
        textColor=colors.HexColor('#0057A8'),
        spaceBefore=15,
        spaceAfter=8
    )
    
    body_style = ParagraphStyle(
        name='JSPLBody',
        fontName='Helvetica',
        fontSize=9,
        leading=12
    )
    
    table_hdr_style = ParagraphStyle(
        name='JSPLTableHdr',
        fontName='Helvetica-Bold',
        fontSize=8,
        leading=10,
        textColor=colors.white
    )
    
    table_cell_style = ParagraphStyle(
        name='JSPLTableCell',
        fontName='Helvetica',
        fontSize=8,
        leading=10
    )
    
    table_cell_mono = ParagraphStyle(
        name='JSPLTableCellMono',
        fontName='Courier',
        fontSize=8,
        leading=10
    )

    story = []
    
    months_map = dict([
        (1, 'Jan'), (2, 'Feb'), (3, 'Mar'), (4, 'Apr'),
        (5, 'May'), (6, 'Jun'), (7, 'Jul'), (8, 'Aug'),
        (9, 'Sep'), (10, 'Oct'), (11, 'Nov'), (12, 'Dec')
    ])
    if filter_type == 'range':
        period_label = f"{months_map.get(from_month)} {from_year} - {months_map.get(to_month)} {to_year}"
    else:
        months_full = [
            (1, 'January'), (2, 'February'), (3, 'March'), (4, 'April'),
            (5, 'May'), (6, 'June'), (7, 'July'), (8, 'August'),
            (9, 'September'), (10, 'October'), (11, 'November'), (12, 'December')
        ]
        period_label = f"{dict(months_full).get(from_month)} {from_year}"

    # JSPL Header Block
    story.append(Paragraph("JINDAL STEEL & POWER LIMITED", title_style))
    story.append(Paragraph(f"Monthly TPM Performance Report — {dept.name} ({dept.code})", subtitle_style))
    story.append(Paragraph(f"Report Period: {period_label}", subtitle_style))
    story.append(Spacer(1, 15))
    
    # Overview Summary
    story.append(Paragraph("Plant Overview Status", header_style))
    overview_data = [
        [Paragraph("Pillar", table_hdr_style),
         Paragraph("Total KPIs", table_hdr_style),
         Paragraph("On Track", table_hdr_style),
         Paragraph("At Risk", table_hdr_style),
         Paragraph("Behind", table_hdr_style)]
    ]
    
    pillars_meta = [
        ('KK', 'Kobetsu Kaizen'),
        ('JH', 'Jishu Hozen'),
        ('PM', 'Planned Maintenance'),
        ('QM', 'Quality Maintenance'),
        ('ET', 'Education & Training'),
        ('DM', 'Initial Flow Control / Design & Management'),
        ('SHE', 'Safety, Health & Environment'),
        ('OTPM', 'Office TPM')
    ]
    
    for code, name in pillars_meta:
        definitions = KPI_DEFINITIONS.get(code, [])
        total = len(definitions)
        on_track = 0
        at_risk = 0
        behind = 0
        has_any_data = False
        
        for d in definitions:
            kpi_values = KPIValue.objects.filter(
                get_date_range_q(prefix='pillar_entry__', from_month=from_month, from_year=from_year, to_month=to_month, to_year=to_year),
                pillar_entry__department=dept,
                pillar_entry__pillar=code,
                sl_no=d['sl_no']
            )
            if kpi_values.exists():
                has_any_data = True
                actual, target, benchmark = aggregate_kpi_actual(kpi_values, d['uom'], d['name'])
                if actual is not None and target is not None:
                    ach = compute_achievement(actual, target, d['name'])
                    if ach >= 90:
                        on_track += 1
                    elif ach >= 75:
                        at_risk += 1
                    else:
                        behind += 1
                        
        overview_data.append([
            Paragraph(f"<b>{code}</b> — {name}", table_cell_style),
            Paragraph(str(total) if has_any_data else "N/A", table_cell_style),
            Paragraph(str(on_track) if has_any_data else "N/A", table_cell_style),
            Paragraph(str(at_risk) if has_any_data else "N/A", table_cell_style),
            Paragraph(str(behind) if has_any_data else "N/A", table_cell_style),
        ])
        
    overview_table = Table(overview_data, colWidths=[230, 70, 70, 70, 70])
    overview_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#003478')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D1DCF0')),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    
    story.append(overview_table)
    story.append(Spacer(1, 20))
    
    # Detailed Pillar KPI Pages
    for code, name in pillars_meta:
        definitions = KPI_DEFINITIONS.get(code, [])
        
        story.append(Paragraph(f"{code} — {name}", header_style))
        
        table_data = [
            [Paragraph("Sl", table_hdr_style),
             Paragraph("KPI Name (UOM)", table_hdr_style),
             Paragraph("Bench", table_hdr_style),
             Paragraph("Target", table_hdr_style),
             Paragraph("Actual", table_hdr_style),
             Paragraph("Achievement", table_hdr_style),
             Paragraph("Remarks", table_hdr_style)]
        ]
        
        for d in definitions:
            kpi_values = KPIValue.objects.filter(
                get_date_range_q(prefix='pillar_entry__', from_month=from_month, from_year=from_year, to_month=to_month, to_year=to_year),
                pillar_entry__department=dept,
                pillar_entry__pillar=code,
                sl_no=d['sl_no']
            )
            
            if kpi_values.exists():
                actual, target, benchmark = aggregate_kpi_actual(kpi_values, d['uom'], d['name'])
                remarks_list = [v.remarks.strip() for v in kpi_values if v.remarks.strip()]
                remarks = " | ".join(remarks_list) if remarks_list else ""
            else:
                actual = None
                target = d['target']
                benchmark = d['benchmark']
                remarks = ''
                
            achievement_str = "—"
            if actual is not None and target is not None:
                ach = compute_achievement(actual, target, d['name'])
                achievement_str = f"{round(ach, 1)}%"
                
            table_data.append([
                Paragraph(d['sl_no'], table_cell_mono),
                Paragraph(f"{d['name']} ({d['uom']})", table_cell_style),
                Paragraph(str(benchmark) if benchmark is not None else "—", table_cell_mono),
                Paragraph(str(target) if target is not None else "—", table_cell_mono),
                Paragraph(str(actual) if actual is not None else "—", table_cell_mono),
                Paragraph(achievement_str, table_cell_mono),
                Paragraph(remarks, table_cell_style)
            ])
            
        pillar_table = Table(table_data, colWidths=[25, 205, 45, 45, 45, 65, 80])
        pillar_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0057A8')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D1DCF0')),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ]))
        story.append(pillar_table)
        story.append(Spacer(1, 15))
        
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()
